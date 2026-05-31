#!/usr/bin/env python3
"""
Vorverarbeitung Parzellen-Stammdaten für die Gartenbegehungs-App.

Joint die Mitgliederliste (Pächter/Kontakt/Adresse) mit den Parzellenflächen (m²)
über den Schlüssel Anl. + Ga-Nr + Ind. und schreibt eine saubere data/parzellen.csv.
Read-only auf den Quell-Dateien. Gibt nur Aggregat-Statistik + anonymisierte Samples aus.
"""
import openpyxl, csv, datetime, os, re
from collections import Counter

MITGL  = "/Users/saschatheissen/Downloads/April 2026 Mitgliederliste.xlsx"
FLAECHE = "/Users/saschatheissen/Library/CloudStorage/Dropbox/Doks/Eigene Dateien Sascha/Coding/Gartenfreunde_old/Daten/Echte Daten/PArzellenfläche.xlsx"
OUT_DIR = "/Users/saschatheissen/Library/CloudStorage/Dropbox/Doks/Eigene Dateien Sascha/Coding/2026_Garftenbegehungsapp/data"
OUT = os.path.join(OUT_DIR, "parzellen.csv")

ANLAGE = {"K": "Kühwasen", "S": "Silberwald"}


def norm(s):
    return re.sub(r"\s+", " ", str(s).strip().lower()) if s is not None else ""


def resolve_sheet(wb, want):
    if want in wb.sheetnames:
        return want
    w = norm(want)
    for n in wb.sheetnames:
        if w in norm(n):
            return n
    return wb.sheetnames[0]


def load_sheet(path, sheet_want):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = resolve_sheet(wb, sheet_want)
    ws = wb[name]
    all_rows = list(ws.iter_rows(values_only=True))
    headers = [(i, norm(v)) for i, v in enumerate(all_rows[0])]
    return name, headers, all_rows[1:]


def find_col(headers, *cands):
    for cand in cands:                     # exact normalized match first
        c = norm(cand)
        for idx, h in headers:
            if h == c:
                return idx
    for cand in cands:                     # then substring
        c = norm(cand)
        for idx, h in headers:
            if c and c in h:
                return idx
    return None


def getv(row, idx):
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return v.strip() if isinstance(v, str) else v


def excel_date(v):
    if v in ("", None):
        return ""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, (int, float)):
        try:
            return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))).isoformat()
        except Exception:
            return str(v)
    return str(v).strip()


def key_of(row, c_anl, c_nr, c_ind):
    anl = str(getv(row, c_anl)).strip().upper()
    nr = getv(row, c_nr)
    if not anl or nr in ("", None):
        return None
    try:
        nr = int(nr)
    except Exception:
        return None
    ind = str(getv(row, c_ind)).strip().lower()
    if ind in ("none", "-"):
        ind = ""
    return (anl, nr, ind)


# --- Mitgliederliste ---
m_sheet, mh, mdata = load_sheet(MITGL, "Gartenverz. & Mitgliederstatus")
c_anl = find_col(mh, "Anl.", "Anl", "Anlage")
c_nr = find_col(mh, "Ga-Nr", "Ga Nr", "GaNr", "Garten-Nr", "Nr")
c_ind = find_col(mh, "Ind.", "Ind", "Index")
c_name = find_col(mh, "Name", "Nachname")
c_vor = find_col(mh, "Vorname")
c_ein = find_col(mh, "Eintritt")
c_str = find_col(mh, "Straße", "Strasse")
c_plz = find_col(mh, "PLZ")
c_ort = find_col(mh, "Wohnort", "Ort")
c_tel = find_col(mh, "Telefon")
c_mob = find_col(mh, "Mobil")
c_mail = find_col(mh, "e-mail", "email", "e mail", "mail")
print(f"Mitglieder-Sheet: {m_sheet!r}")
print("  cols:", dict(anl=c_anl, nr=c_nr, ind=c_ind, name=c_name, vor=c_vor,
                       ein=c_ein, str=c_str, plz=c_plz, ort=c_ort, tel=c_tel, mob=c_mob, mail=c_mail))

members, anl_seen = {}, set()
for row in mdata:
    k = key_of(row, c_anl, c_nr, c_ind)
    if not k:
        continue
    anl_seen.add(k[0])
    members[k] = {
        "nachname": getv(row, c_name), "vorname": getv(row, c_vor),
        "email": getv(row, c_mail), "telefon": getv(row, c_mob) or getv(row, c_tel),
        "strasse": getv(row, c_str), "plz": getv(row, c_plz), "ort": getv(row, c_ort),
        "eintritt": excel_date(getv(row, c_ein)),
    }

# --- Parzellenfläche ---
f_sheet, fh, fdata = load_sheet(FLAECHE, "JR 2024-2025")
f_anl = find_col(fh, "Anl.", "Anl", "Anlage")
f_nr = find_col(fh, "Ga-Nr", "Ga Nr", "GaNr", "Nr")
f_ind = find_col(fh, "Ind.", "Ind", "Index")
f_m2 = find_col(fh, "parzellengroesse", "parzellengröße", "parzellengrösse")
f_ar = find_col(fh, "Parzellenfläche", "Parzellenflaeche")
print(f"Fläche-Sheet: {f_sheet!r}")
print("  cols:", dict(anl=f_anl, nr=f_nr, ind=f_ind, m2=f_m2, ar=f_ar))

sizes = {}
for row in fdata:
    k = key_of(row, f_anl, f_nr, f_ind)
    if not k:
        continue
    m2 = getv(row, f_m2)
    if m2 in ("", None):
        ar = getv(row, f_ar)
        try:
            m2 = round(float(ar) * 100)
        except Exception:
            m2 = ""
    else:
        try:
            m2 = round(float(m2))
        except Exception:
            m2 = ""
    sizes[k] = m2

# --- Merge ---
cols = ["parzelle_id", "anlage_kuerzel", "anlage", "nummer", "index",
        "nachname", "vorname", "email", "telefon", "strasse", "plz", "ort",
        "eintritt", "groesse_m2", "soll_gemuese_m2",
        "in_mitglieder", "hat_flaeche"]
rows_out = []
for (anl, nr, ind) in sorted(set(members) | set(sizes)):
    if anl not in ANLAGE:          # nur echte Anlagen (K, S); MK u. a. ignorieren
        continue
    m = members.get((anl, nr, ind), {})
    size = sizes.get((anl, nr, ind), "")
    soll = round(size / 6) if isinstance(size, (int, float)) and size else ""
    rows_out.append({
        "parzelle_id": f"{anl}{nr}{ind}", "anlage_kuerzel": anl, "anlage": ANLAGE.get(anl, "?"),
        "nummer": nr, "index": ind,
        "nachname": m.get("nachname", ""), "vorname": m.get("vorname", ""),
        "email": m.get("email", ""), "telefon": m.get("telefon", ""),
        "strasse": m.get("strasse", ""), "plz": m.get("plz", ""), "ort": m.get("ort", ""),
        "eintritt": m.get("eintritt", ""),
        "groesse_m2": size, "soll_gemuese_m2": soll,
        "in_mitglieder": (anl, nr, ind) in members, "hat_flaeche": (anl, nr, ind) in sizes,
    })

os.makedirs(OUT_DIR, exist_ok=True)
with open(OUT, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=cols)
    w.writeheader()
    w.writerows(rows_out)

# --- QA (ohne personenbezogene Daten) ---
by_anl = Counter(r["anlage_kuerzel"] for r in rows_out)
missing_size = [r["parzelle_id"] for r in rows_out if not r["groesse_m2"]]
only_members = [r["parzelle_id"] for r in rows_out if r["in_mitglieder"] and not r["hat_flaeche"]]
only_size = [r["parzelle_id"] for r in rows_out if r["hat_flaeche"] and not r["in_mitglieder"]]
no_name = [r["parzelle_id"] for r in rows_out if not (r["nachname"] or r["vorname"])]
print("\n=== ERGEBNIS ===")
print("Datei:", OUT)
print("Parzellen gesamt:", len(rows_out))
print("nach Anlage:", dict(by_anl))
print("unbekannte Anl.-Kürzel:", sorted(anl_seen - set(ANLAGE)))
print("ohne Fläche:", len(missing_size), "->", missing_size[:15])
print("nur Mitglieder, keine Fläche:", len(only_members), "->", only_members[:15])
print("nur Fläche, kein Mitglied:", len(only_size), "->", only_size[:15])
print("ohne Pächternamen (frei/leer?):", len(no_name), "->", no_name[:15])
print("\nSPALTEN:", cols)
print("SAMPLE (anonymisiert):")
priv = {"nachname", "vorname", "email", "telefon", "strasse", "plz", "ort"}
for r in rows_out[:3]:
    print({k: ("<gesetzt>" if (k in priv and r[k]) else ("" if k in priv else r[k])) for k in cols})
